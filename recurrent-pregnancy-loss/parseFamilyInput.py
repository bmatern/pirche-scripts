import argparse
import datetime
from os.path import isdir, join
from os import makedirs
from openpyxl import load_workbook

def getTwoFieldTyping(cell=None, locus=None):
    print('\nCell:' + str(cell))
    cellType=cell.data_type
    print('Type:' + str(cellType))
    print('Value:' + str(cell.value))


    if cell.is_date:

        #print('Hours:' + str(cell.parts['hour']))
        if isinstance (cell.value,datetime.timedelta):
            print('The cell is Timedelta!')
            totalSeconds=cell.value.total_seconds()
            hours, remainder = divmod(totalSeconds, 3600)
            minutes, seconds = divmod(remainder, 60)
            #print('hours:' + str(int(hours)))
            #print('minutes:' + str(int(minutes)))
            #print('seconds: '+ str(int(seconds)))
            rawTyping = str(int(hours)) + ':' + str(int(minutes)) + ':' + str(int(seconds))
        elif isinstance (cell.value,datetime.time):
            print('The cell is Datetime.time!')
            #print('Formatted:' + cell.value.strftime('%h:%m'))
            rawTyping = cell.value
        else:
            raise Exception('Not a timedelta:' + str(type(cell.value)))
    else:
        rawTyping=cell.value
        #

    print('raw typing:' + str(rawTyping))





    if(rawTyping is None):
        return None
    else:
        nomenclatureTokens = str(rawTyping).replace('.',':').split(':')
        if(len(nomenclatureTokens)>1):
            for tokenIndex, nomenclatureToken in enumerate(nomenclatureTokens):
                if(len(nomenclatureToken)==1):
                    print ('converting (' + nomenclatureToken + ') to (0' + nomenclatureToken + ')' )
                    nomenclatureTokens[tokenIndex] = '0' + nomenclatureTokens[tokenIndex]
            twofieldTyping = str(locus) + '*' + nomenclatureTokens[0] + ':' + nomenclatureTokens[1]
            print('returning ' + str(twofieldTyping))
            return twofieldTyping
        else:
            raise Exception('Cannot find a 2 field typing in this string:' + str(rawTyping))

def parseExcelForFamilyData(outputFileName=None, excelFileName=None, headers=True, newLine='\n', separator=','):
    print('Parsing Excel File:' + str(excelFileName))
    print('Writing output file:' + str(outputFileName))

    with open(outputFileName, 'w') as outputFile:

        excelData = load_workbook(excelFileName)

        idColumn='B'
        a1Column='C'
        a2Column='D'
        b1Column='E'
        b2Column='F'
        c1Column='G'
        c2Column='H'
        drb11Column='I'
        drb12Column='J'
        dqb11Column='K'
        dqb12Column='L'


        firstSheet = excelData[excelData.sheetnames[0]]
        if (headers):
            startRowIndex = 2
        else:
            startRowIndex = 1


        for rowIndexRaw, row in enumerate(firstSheet.iter_rows(min_row=startRowIndex, values_only=True)):
            print('Row:' + str(row))

            if(str(row)=='(None, None, None, None, None, None, None, None, None, None, None, None, None)'):
                #print('That was a None Row')
                outputFile.write(separator + newLine)

            else:
                rowTyping={}
                rowTyping['sampleid'] = firstSheet[idColumn + str(rowIndexRaw + 2)].value
                rowTyping['A1']=getTwoFieldTyping(locus='A', cell=firstSheet[a1Column + str(rowIndexRaw + 2)])
                rowTyping['A2'] = getTwoFieldTyping(locus='A', cell=firstSheet[a2Column + str(rowIndexRaw + 2)])
                rowTyping['B1']=getTwoFieldTyping(locus='B', cell=firstSheet[b1Column + str(rowIndexRaw + 2)])
                rowTyping['B2'] = getTwoFieldTyping(locus='B', cell=firstSheet[b2Column + str(rowIndexRaw + 2)])
                rowTyping['C1']=getTwoFieldTyping(locus='C', cell=firstSheet[c1Column + str(rowIndexRaw + 2)])
                rowTyping['C2'] = getTwoFieldTyping(locus='C', cell=firstSheet[c2Column + str(rowIndexRaw + 2)])
                rowTyping['DRB11']=getTwoFieldTyping(locus='DRB1', cell=firstSheet[drb11Column + str(rowIndexRaw + 2)])
                rowTyping['DRB12'] = getTwoFieldTyping(locus='DRB1', cell=firstSheet[drb12Column + str(rowIndexRaw + 2)])
                rowTyping['DQB11']=getTwoFieldTyping(locus='DQB1', cell=firstSheet[dqb11Column + str(rowIndexRaw + 2)])
                rowTyping['DQB12'] = getTwoFieldTyping(locus='DQB1', cell=firstSheet[dqb12Column + str(rowIndexRaw + 2)])
                outputFile.write(rowTyping['sampleid']
                + separator + rowTyping['A1']
                + separator + rowTyping['A2']
                + separator + rowTyping['B1']
                + separator + rowTyping['B2']
                + separator + rowTyping['C1']
                + separator + rowTyping['C2']
                + separator + rowTyping['DQB11']
                + separator + rowTyping['DQB12']
                + separator + rowTyping['DRB11']
                + separator + rowTyping['DRB12']
                + separator + newLine)






if __name__ == '__main__':

    parser = argparse.ArgumentParser()

    parser.add_argument("-v", "--verbose", help="verbose operation", action="store_true")
    parser.add_argument("-o", "--output", help="output directory", required=True)
    parser.add_argument("-x", "--excel", help="input excel file", required=True)

    args = parser.parse_args()

    verbose = args.verbose
    if verbose:
        print("verbose mode active")

    if (verbose):
        print("Excel:" + args.excel)
        print("Output:" + args.output)

    if(not isdir(args.output)):
        makedirs(args.output)

    outputFileName = join(args.output, 'PircheInputFile.txt')

    familyData = parseExcelForFamilyData(excelFileName=args.excel, outputFileName=outputFileName)

    print('Done.')