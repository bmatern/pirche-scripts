import argparse
import csv
from os import makedirs
from os.path import join, isdir

from openpyxl import load_workbook
from numpy.random import choice, seed

def cleanHLA(locus, allele):
    allele = allele.replace("g", "")
    if "*" in allele:
        allele = allele[allele.index("*") + 1:]
    if ":" not in allele:
        if len(allele) > 3:
            allele = allele[0:2] + ":" + allele [2:4]
        else:
            print("ERROR: cannot clean " + locus + " " + allele)
    return locus + "*" + allele.replace("g", "")


def parseArgs():
    parser = argparse.ArgumentParser()

    parser.add_argument("-hp", "--haplotypes", help="NMDP haplotype table (either 2007 or 2011 or equally formatted)",  type=str, required=True)
    parser.add_argument("-n", "--number", help="number of individuals to simulate", type=int, required=False, default=1000)
    parser.add_argument("-p", "--population", help="population short code as used in the header row", required=True)
    parser.add_argument("-o", "--output", help="output directory", required=True)
    parser.add_argument("-s", "--seed", help="random seed", required=False, type=int, default=12345)
    args = parser.parse_args()
    return args


def loadData(filename=None):
    workbook = load_workbook(filename=filename)
    worksheet = workbook[workbook.sheetnames[0]]

    freq_suffix = "_FREQ"
    rank_suffix = "_RANK"
    col_idx = {}
    for colIndexRaw, col in enumerate(worksheet.iter_cols()):
        if(col[0].value is not None and len(str(col[0].value))>0):
            headerName = str(col[0].value).strip().upper()

            print(str(colIndexRaw) + " -> " + headerName)
            col_idx[headerName] = colIndexRaw

    print(spacer)
    print('col_idx=' + str(col_idx))

    haplotypes = []

    #for row_idx in range(1, worksheet.nrows):
    for rowIndexRaw, row in enumerate(worksheet.iter_rows()):
        #print('Row:' + str(row))
        if(rowIndexRaw >= 1):
            rank=str((row[col_idx[args.population + rank_suffix]].value)).strip().upper()
            if(rank != 'NA'):
                haplotype = {
                    'id': rowIndexRaw,
                    'A':  cleanHLA("A", row[col_idx['A']].value),
                    'B':  cleanHLA("B", row[col_idx["B"]].value),
                    'C':  cleanHLA("C", row[col_idx["C"]].value),
                    'DRB1':  cleanHLA("DRB1", row[col_idx["DRB1"]].value),
                    'DQB1':  cleanHLA("DQB1", row[col_idx["DQB1"]].value),
                    'freq':  float(row[col_idx[args.population + freq_suffix]].value),
                }
                haplotypes.append(haplotype)
                #print('appended haplotype:' + str(haplotype))
            else:
                pass
                #print('row is empty:' + str(row))


    print("read " + str(len(haplotypes)) + " haplotypes")
    print(spacer)

    haplotypes.sort(key=lambda x: x['freq'], reverse=True)

    return haplotypes

def writeGenotypes(haplotypeData=None, genotypeFileName=None, targetGenotypeCount=1):
    weights = [haplotype['freq'] for haplotype in haplotypeData]
    frequencyTotal=sum(weights)

    print("Frequency sum is: " + str(frequencyTotal))
    print(spacer)

    with open(genotypeFileName, "w") as output:
        writer = csv.writer(output, delimiter=',')
        writer.writerow(["genotype", "A1", "A2", "B1", "B2", "C1", "C2", "DRB11", "DRB12", "DQB11", "DQB12"])

        currentGenotypeCount = 0

        while(currentGenotypeCount < targetGenotypeCount):
            left, right = list(choice(haplotypeData, 2, p=weights))

            # Somewhat unlikely, but if we find the dummy haplotype, then don't use that:
            if str(left['id']) == 'FILLER' or str(right['id']) == 'FILLER':
                #print('Skipping the filler haplotype... This should be rare')
                continue

            writer.writerow(['GENOTYPE_' + str(left['id']) + "_" + str(right['id']), left['A'], right['A'], left['B'], right['B'], left['C'], right['C'], left['DRB1'], right['DRB1'], left['DQB1'], right['DQB1']])

            currentGenotypeCount += 1

    print(spacer)
    print("created " + str(args.number) + " genotypes.")
    print(spacer)
    print("successfully stored genotypes in file:" + genotypeFileName)
    print(spacer)

def addDummyFiller(haplotypeData=None):
    weights = [haplotype['freq'] for haplotype in haplotypeData]
    totalFrequency = sum(weights)

    # Weights should sum to 1.0. They don't....I could scale all probabilities, however there are difficulties in floating point calculations.
    #   Silly solution: put in a dummy haplotype that is not actually used.
    #   These dummys represent the "missing" members of the population based on frequency cutoffs.
    missingFrequency = 1 - totalFrequency

    dummyHaplotype = {}
    dummyHaplotype['id'] = 'FILLER'
    dummyHaplotype['freq'] = missingFrequency
    haplotypeData.append(dummyHaplotype)

    weights = [haplotype['freq'] for haplotype in haplotypeData]
    totalFrequency = sum(weights)
    if totalFrequency != 1:
        raise Exception('Total frequency is not 1.')

    # Now they sum to 1.
    return haplotypeData


def writeHaplotypes(haplotypeData=None, haplotypeFileName=None, targetHaplotypeCount=None):
    weights = [haplotype['freq'] for haplotype in haplotypeData]
    frequencyTotal = sum(weights)

    print("Frequency sum is: " + str(frequencyTotal))
    print(spacer)

    with open(haplotypeFileName, "w") as output:
        writer = csv.writer(output, delimiter=',')
        writer.writerow(["haplotype", "A", "B", "C", "DRB1", "DQB1"])

        currentHaplotypeCount = 0

        while (currentHaplotypeCount < targetHaplotypeCount):
            haplotype = list(choice(haplotypeData, 1, p=weights))[0]

            # Somewhat unlikely, but if we find the dummy haplotype, then don't use that:
            if str(haplotype['id']) == 'FILLER':
                # print('Skipping the filler haplotype... This should be rare')
                continue

            writer.writerow(['HAPLOTYPE_' + str(haplotype['id']) , haplotype['A'], haplotype['B'], haplotype['C'], haplotype['DRB1'], haplotype['DQB1']])
            currentHaplotypeCount += 1

    print(spacer)
    print("created " + str(args.number) + " Haplotypes.")
    print(spacer)
    print("successfully stored Haplotypes in file:" + haplotypeFileName)
    print(spacer)


if __name__ == '__main__':

    args = parseArgs()

    # read files
    spacer = '-'*40
    print("read haplotype file:" + str(args.haplotypes))
    print(spacer)

    haplotypeData = loadData(filename=args.haplotypes)
    haplotypeData = addDummyFiller(haplotypeData=haplotypeData)

    if not isdir(args.output):
        makedirs(args.output)

    seed(args.seed)

    genotypeFileName = join(args.output, 'Genotypes.csv')
    writeGenotypes(haplotypeData=haplotypeData, genotypeFileName=genotypeFileName, targetGenotypeCount=args.number)

    haplotypeFileName = join(args.output, 'Haplotypes.csv')
    writeHaplotypes(haplotypeData=haplotypeData, haplotypeFileName=haplotypeFileName, targetHaplotypeCount=args.number)
