import argparse
import datetime
from os.path import isdir, join
from os import makedirs
from openpyxl import load_workbook


def getTwoFieldTyping(cell=None, locus=None):
    #print('\nCell:' + str(cell))
    cellType=cell.data_type
    #print('Type:' + str(cellType))
    #print('Value:' + str(cell.value))


    if cell.is_date:
        print('WARNING It is a date! Make sure this works right! (no problem if this is None):' + str(cell.value))
        if(cell.value=='None' or cell.value is None):
            return None

        if isinstance (cell.value,datetime.timedelta):
            print('The cell is Timedelta!:' + str(cell.value))
            totalSeconds=cell.value.total_seconds()
            hours, remainder = divmod(totalSeconds, 3600)
            minutes, seconds = divmod(remainder, 60)
            #print('hours:' + str(int(hours)))
            #print('minutes:' + str(int(minutes)))
            #print('seconds: '+ str(int(seconds)))
            rawTyping = str(int(hours)) + ':' + str(int(minutes)) + ':' + str(int(seconds))
        elif isinstance (cell.value,datetime.time):
            print('The cell is Datetime.time!')
            #print('Formatted:' + cell.value.strftime('%h:%m'))
            rawTyping = cell.value
        else:
            raise Exception('Not a timedelta:' + str(type(cell.value)))
    else:
        rawTyping=str(cell.value).strip()
        #

    #print('raw typing:' + str(rawTyping))





    if(rawTyping is None or rawTyping=='None' or len(rawTyping)==0):
        return None
    else:
        #print('Interpreting raw typing ' + rawTyping)
        nomenclatureTokens = str(rawTyping).replace('.',':').split(':')
        if(len(nomenclatureTokens)==1):

            if(len(rawTyping)==2 and isAllDigits(rawTyping)):
                # 1-field typing, allele group only
                twofieldTyping = str(locus) + '*' + rawTyping[0:2]
                return twofieldTyping
            elif(len(rawTyping)==4 and isAllDigits(rawTyping)):
                # 2-field typing missing the : delimiter
                twofieldTyping = str(locus) + '*' + rawTyping[0:2] + ':' + rawTyping[2:4]
                return twofieldTyping
            elif(len(rawTyping)==4 and isAllDigits(rawTyping[0:2]) and rawTyping[2:4]=='XX'):
                #1 field allele call with XX at the end. This is just a 1 field allele.
                twofieldTyping = str(locus) + '*' + rawTyping[0:2]
                #print(rawTyping + ' is a 1-field allele ending in XX: ' + str(rawTyping) + ' which is converted to ' + str(twofieldTyping))
                return twofieldTyping
            elif(len(rawTyping)==5 and isAllDigits(rawTyping[0:4]) and not(isAllDigits(rawTyping[4:]))):
                # 2 field typing with expression marker
                #print('This is an allele with expression marker:' + str(rawTyping))
                twofieldTyping = str(locus) + '*' + rawTyping[0:2] + ':' + rawTyping[2:4] + rawTyping[4:]
                #print('twoFieldTypingWithExpressionmarker=' + str(twofieldTyping))
                return twofieldTyping
            elif(len(rawTyping)>3 and isAllDigits(rawTyping[0:2]) and not(isAllDigits(rawTyping[2:]))):
                # 1-field typing with probably a MAC Code afterwards?
                twofieldTyping = str(locus) + '*' + rawTyping[0:2] + ':' + rawTyping[2:]
                #print('This looks like a 1-field allele with MAC Codes:' + str(rawTyping) + ' converted to ' + str(twofieldTyping))
                return twofieldTyping
            elif(len(rawTyping)>3 and isAllDigits(rawTyping[0:1]) and not(isAllDigits(rawTyping[1:]))):
                # 1-field typing with probably a MAC Code afterwards?
                # This is a single digit, followed by some text. 1-field typing with a Mac code, plus an excel conversion to get rid of a leading 0.
                # ex. 7ANVB
                twofieldTyping = str(locus) + '*0' + rawTyping[0:1] + ':' + rawTyping[1:]
                print('This looks like a 1-field allele with MAC Codes:' + str(rawTyping) + ' converted to ' + str(twofieldTyping))
                return twofieldTyping

            else:
                raise Exception('What to do with this raw typing?' + str(rawTyping))



        elif(len(nomenclatureTokens)>1):
            # TODO: Not handling NMDP codes here. Keep them as 1-field alleles for now?
            for tokenIndex, nomenclatureToken in enumerate(nomenclatureTokens):
                if(len(nomenclatureToken)==1):
                    print ('converting (' + nomenclatureToken + ') to (0' + nomenclatureToken + ')' )
                    nomenclatureTokens[tokenIndex] = '0' + nomenclatureTokens[tokenIndex]
            twofieldTyping = str(locus) + '*' + nomenclatureTokens[0] + ':' + nomenclatureTokens[1]
            #print('returning ' + str(twofieldTyping))
            return twofieldTyping
        else:
            raise Exception('Cannot find a 2 field typing in this string:' + str(rawTyping))


def isAllDigits(inputText=None):
    # Can i just try to convert it to an integer? Maybe that's good enough....
    try:
        asInteger = int(inputText)
        return True
    except Exception as e:
        return False
    return False





def readExcelFile(excelFilename=None, tabNumber=None, headers=True, lastRowIndex=9999):
    patientTypings={}
    donorTypings={}


    excelData = load_workbook(excelFilename)

    idColumn = 'A'
    pA1Column = 'B'
    pA2Column = 'C'
    pB1Column = 'D'
    pB2Column = 'E'
    pC1Column = 'F'
    pC2Column = 'G'
    pDrb11Column = 'H'
    pDrb12Column = 'I'
    pDqb11Column = 'J'
    pDqb12Column = 'K'

    dA1Column = 'L'
    dA2Column = 'M'
    dB1Column = 'N'
    dB2Column = 'O'
    dC1Column = 'P'
    dC2Column = 'Q'
    dDrb11Column = 'R'
    dDrb12Column = 'S'
    dDqb11Column = 'T'
    dDqb12Column = 'U'

    # input tab # is 1-based
    dataSheet = excelData[excelData.sheetnames[int(tabNumber)-1]]

    if (headers):
        startRowIndex = 2
    else:
        startRowIndex = 1

    for rowIndexRaw, row in enumerate(dataSheet.iter_rows(min_row=startRowIndex, max_row=lastRowIndex, values_only=True)):
        #print('Row:' + str(row))

        # Trim out nonsense, if there is nothing left than this was an empty cell.
        if (str(row).replace('None', '').replace('(', '').replace(')', '').replace(',', '').replace('\'', '').replace(' ', '') == ''):
            pass
            #print('row # ' + str(rowIndexRaw + 2) + ' is empty.')

        else:
            # TODO: might not be an int??
            sampleId = int(str(dataSheet[idColumn + str(rowIndexRaw + 2)].value).strip())
            #print('sample:' + str(sampleId))
            patientTyping = {}
            patientTyping['A1'] = getTwoFieldTyping(locus='A', cell=dataSheet[pA1Column + str(rowIndexRaw + 2)])
            patientTyping['A2'] = getTwoFieldTyping(locus='A', cell=dataSheet[pA2Column + str(rowIndexRaw + 2)])
            patientTyping['B1'] = getTwoFieldTyping(locus='B', cell=dataSheet[pB1Column + str(rowIndexRaw + 2)])
            patientTyping['B2'] = getTwoFieldTyping(locus='B', cell=dataSheet[pB2Column + str(rowIndexRaw + 2)])
            patientTyping['C1'] = getTwoFieldTyping(locus='C', cell=dataSheet[pC1Column + str(rowIndexRaw + 2)])
            patientTyping['C2'] = getTwoFieldTyping(locus='C', cell=dataSheet[pC2Column + str(rowIndexRaw + 2)])
            patientTyping['DRB11'] = getTwoFieldTyping(locus='DRB1', cell=dataSheet[pDrb11Column + str(rowIndexRaw + 2)])
            patientTyping['DRB12'] = getTwoFieldTyping(locus='DRB1', cell=dataSheet[pDrb12Column + str(rowIndexRaw + 2)])
            patientTyping['DQB11'] = getTwoFieldTyping(locus='DQB1', cell=dataSheet[pDqb11Column + str(rowIndexRaw + 2)])
            patientTyping['DQB12'] = getTwoFieldTyping(locus='DQB1', cell=dataSheet[pDqb12Column + str(rowIndexRaw + 2)])
            patientTypings[sampleId] = patientTyping

            donorTyping = {}
            donorTyping['A1'] = getTwoFieldTyping(locus='A', cell=dataSheet[dA1Column + str(rowIndexRaw + 2)])
            donorTyping['A2'] = getTwoFieldTyping(locus='A', cell=dataSheet[dA2Column + str(rowIndexRaw + 2)])
            donorTyping['B1'] = getTwoFieldTyping(locus='B', cell=dataSheet[dB1Column + str(rowIndexRaw + 2)])
            donorTyping['B2'] = getTwoFieldTyping(locus='B', cell=dataSheet[dB2Column + str(rowIndexRaw + 2)])
            donorTyping['C1'] = getTwoFieldTyping(locus='C', cell=dataSheet[dC1Column + str(rowIndexRaw + 2)])
            donorTyping['C2'] = getTwoFieldTyping(locus='C', cell=dataSheet[dC2Column + str(rowIndexRaw + 2)])
            donorTyping['DRB11'] = getTwoFieldTyping(locus='DRB1', cell=dataSheet[dDrb11Column + str(rowIndexRaw + 2)])
            donorTyping['DRB12'] = getTwoFieldTyping(locus='DRB1', cell=dataSheet[dDrb12Column + str(rowIndexRaw + 2)])
            donorTyping['DQB11'] = getTwoFieldTyping(locus='DQB1', cell=dataSheet[dDqb11Column + str(rowIndexRaw + 2)])
            donorTyping['DQB12'] = getTwoFieldTyping(locus='DQB1', cell=dataSheet[dDqb12Column + str(rowIndexRaw + 2)])
            donorTypings[sampleId] = donorTyping

    return patientTypings, donorTypings


def getTypingIfExists(typings=None, locus=None, delimiter=None):
    if(locus in typings.keys()):
        if typings[locus] is None:
            return ''
        else:
            return typings[locus] + delimiter
    else:
        return ''


def writePircheFile(patientsPerID=None, donorsPerID=None, outputFilename=None, delimiter = ',' ,newline='\r\n' ):
    with open(outputFilename, 'w') as outputFile:
        for index, sampleID in enumerate(list(patientsPerID.keys())):
            print('writing sample id ' + str(sampleID))
            # write patient types
            patientLine = 'P' + str(sampleID) + delimiter
            patientLine = patientLine + getTypingIfExists(typings=patientsPerID[sampleID], locus='A1', delimiter=delimiter)
            patientLine = patientLine + getTypingIfExists(typings=patientsPerID[sampleID], locus='A2', delimiter=delimiter)
            patientLine = patientLine + getTypingIfExists(typings=patientsPerID[sampleID], locus='B1', delimiter=delimiter)
            patientLine = patientLine + getTypingIfExists(typings=patientsPerID[sampleID], locus='B2', delimiter=delimiter)
            patientLine = patientLine + getTypingIfExists(typings=patientsPerID[sampleID], locus='C1', delimiter=delimiter)
            patientLine = patientLine + getTypingIfExists(typings=patientsPerID[sampleID], locus='C2', delimiter=delimiter)
            patientLine = patientLine + getTypingIfExists(typings=patientsPerID[sampleID], locus='DRB11', delimiter=delimiter)
            patientLine = patientLine + getTypingIfExists(typings=patientsPerID[sampleID], locus='DRB12', delimiter=delimiter)
            patientLine = patientLine + getTypingIfExists(typings=patientsPerID[sampleID], locus='DQB11', delimiter=delimiter)
            patientLine = patientLine + getTypingIfExists(typings=patientsPerID[sampleID], locus='DQB12', delimiter=delimiter)
            # trailing comma
            patientLine = patientLine.strip(delimiter) + newline
            outputFile.write(patientLine)


            # write CB types
            donorLine = 'CB' + str(sampleID) + delimiter
            donorLine = donorLine + getTypingIfExists(typings=donorsPerID[sampleID], locus='A1', delimiter=delimiter)
            donorLine = donorLine + getTypingIfExists(typings=donorsPerID[sampleID], locus='A2', delimiter=delimiter)
            donorLine = donorLine + getTypingIfExists(typings=donorsPerID[sampleID], locus='B1', delimiter=delimiter)
            donorLine = donorLine + getTypingIfExists(typings=donorsPerID[sampleID], locus='B2', delimiter=delimiter)
            donorLine = donorLine + getTypingIfExists(typings=donorsPerID[sampleID], locus='C1', delimiter=delimiter)
            donorLine = donorLine + getTypingIfExists(typings=donorsPerID[sampleID], locus='C2', delimiter=delimiter)
            donorLine = donorLine + getTypingIfExists(typings=donorsPerID[sampleID], locus='DRB11', delimiter=delimiter)
            donorLine = donorLine + getTypingIfExists(typings=donorsPerID[sampleID], locus='DRB12', delimiter=delimiter)
            donorLine = donorLine + getTypingIfExists(typings=donorsPerID[sampleID], locus='DQB11', delimiter=delimiter)
            donorLine = donorLine + getTypingIfExists(typings=donorsPerID[sampleID], locus='DQB12', delimiter=delimiter)
            # trailing comma
            donorLine = donorLine.strip(delimiter) + newline
            outputFile.write(donorLine)


            # Separate with another comma
            if(index < len(list(patientsPerID.keys()))-1):
                outputFile.write(delimiter + newline)


if __name__ == '__main__':

    parser = argparse.ArgumentParser()
    parser.add_argument("-v", "--verbose", help="verbose operation", action="store_true")
    parser.add_argument("-o", "--output", help="output directory", required=True)
    parser.add_argument("-x", "--excel", help="input excel file", required=True)
    parser.add_argument("-t", "--tab", help="Tab (1-based) in excel to pull data from")

    args = parser.parse_args()

    verbose = args.verbose
    if verbose:
        print("verbose mode active")

    if (verbose):
        print("Excel:" + args.excel)
        print("Looking at (1-based) tab# " + str(args.tab))
        print("Output:" + args.output)

    if(not isdir(args.output)):
        makedirs(args.output)

    print('Processing Reading this excel file:' + str(args.excel))
    patientsPerID, donorsPerID = readExcelFile(excelFilename=args.excel, tabNumber=args.tab)
    print('patientsPerID:' + str(patientsPerID))

    outputFilename = join(args.output, 'pirche_input_file.csv')
    print('Writing output file:' + str(outputFilename))
    writePircheFile(patientsPerID=patientsPerID, donorsPerID=donorsPerID, outputFilename=outputFilename)

    print('Done.')